from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from mocktrade.models import Cosmetic, Hero, MarketRecord, MockTrading, UserAccount


HERO_TYPES = {
    "CK": "Strength",
    "DK": "Strength",
    "先知": "Intelligence",
    "冰女": "Intelligence",
    "剧毒": "Intelligence",
    "大牛": "Strength",
    "天怒": "Intelligence",
    "女王": "Intelligence",
    "小Y": "Intelligence",
    "小小": "Strength",
    "小牛": "Strength",
    "小鱼": "Agility",
    "小鹿": "Intelligence",
    "小黑": "Agility",
    "尸王": "Strength",
    "屠夫": "Strength",
    "拉比克": "Intelligence",
    "拍拍": "Agility",
    "斧王": "Strength",
    "斯温": "Strength",
    "水人": "Agility",
    "海民": "Strength",
    "潮汐": "Strength",
    "火女": "Intelligence",
    "火猫": "Agility",
    "电棍": "Agility",
    "蓝猫": "Intelligence",
    "谜团": "Intelligence",
    "酒仙": "Strength",
    "龙骑": "Strength",
}

HOLDINGS_SHEET_INDEX = 0
DATABASE_SHEET_INDEX = 2


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return datetime.strptime(str(int(value)), "%Y%m%d").date()
    if isinstance(value, str):
        return datetime.strptime(value.strip(), "%Y%m%d").date()
    raise ValueError(f"Unsupported date value: {value!r}")


def resolve_workbook_path(path_option):
    base_dir = Path(settings.BASE_DIR)

    if path_option:
        workbook_path = Path(path_option)
        if not workbook_path.is_absolute():
            workbook_path = base_dir / workbook_path
        if not workbook_path.exists():
            raise CommandError(f"Workbook not found: {workbook_path}")
        return workbook_path

    preferred = base_dir / "dota2饰品.xlsx"
    if preferred.exists():
        return preferred

    candidates = sorted(base_dir.glob("*.xlsx"), key=lambda item: (-item.stat().st_size, item.name))
    if not candidates:
        raise CommandError("No .xlsx workbook found in the project root.")
    return candidates[0]


def parse_database_sheet(worksheet):
    header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    date_columns = []

    for index in range(2, len(header), 2):
        raw_date = header[index]
        if raw_date is None:
            break
        date_columns.append((index, parse_excel_date(raw_date)))

    raw_rows = []
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        hero_name = clean_text(row[0])
        slot_name = clean_text(row[1])
        if not hero_name or not slot_name:
            continue
        raw_rows.append((hero_name, slot_name, row))

    hero_counts = Counter(hero_name for hero_name, _, _ in raw_rows)
    item_specs = []

    for hero_name, slot_name, row in raw_rows:
        item_name = hero_name if hero_counts[hero_name] == 1 else f"{hero_name}-{slot_name}"
        market_points = []

        for index, point_date in date_columns:
            price = row[index] if index < len(row) else None
            quantity = row[index + 1] if index + 1 < len(row) else None
            if price is None or quantity is None:
                continue
            market_points.append(
                {
                    "date": point_date,
                    "price": Decimal(str(price)),
                    "quantity": int(quantity),
                }
            )

        item_specs.append(
            {
                "hero_name": hero_name,
                "slot_name": slot_name,
                "item_name": item_name,
                "market_points": market_points,
            }
        )

    return item_specs


def parse_holdings_sheet(worksheet):
    holdings = []
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        hero_name = clean_text(row[1] if len(row) > 1 else None)
        if not hero_name:
            continue

        current_price = row[5] if len(row) > 5 else None
        buy_date = row[3] if len(row) > 3 else None
        if buy_date is None or current_price is None:
            continue
        holdings.append(
            {
                "hero_name": hero_name,
                "current_price": Decimal(str(current_price)) if current_price is not None else None,
                "buy_date": buy_date.date() if isinstance(buy_date, datetime) else buy_date,
            }
        )

    return holdings


def ensure_auth_user(username, password):
    user_model = get_user_model()
    auth_user = user_model.objects.filter(username=username).first()
    if auth_user:
        return auth_user, False

    if not password:
        raise CommandError(
            f"Auth user '{username}' does not exist. Re-run with --password to create it."
        )

    return user_model.objects.create_user(username=username, password=password), True


def ensure_trading_account(username):
    user_account, user_account_created = UserAccount.objects.get_or_create(
        username=username,
        defaults={"password": "", "status": True},
    )
    trading, trading_created = MockTrading.objects.get_or_create(
        user=user_account,
        defaults={"balance": Decimal("10000.00")},
    )
    return user_account, trading, user_account_created, trading_created


def import_market_data(item_specs):
    cosmetics_created = 0
    heroes_created = 0
    records_created = 0
    hero_to_items = defaultdict(list)

    for spec in item_specs:
        cosmetic, created = Cosmetic.objects.get_or_create(item_name=spec["item_name"])
        if created:
            cosmetics_created += 1
        spec["cosmetic"] = cosmetic
        spec["latest_price"] = spec["market_points"][-1]["price"] if spec["market_points"] else None
        hero_to_items[spec["hero_name"]].append(spec)

        for point in spec["market_points"]:
            _, record_created = MarketRecord.objects.update_or_create(
                item=cosmetic,
                date=point["date"],
                defaults={
                    "price": point["price"],
                    "quantity": point["quantity"],
                },
            )
            if record_created:
                records_created += 1

    for hero_name, specs in hero_to_items.items():
        hero, created = Hero.objects.update_or_create(
            hero_name=hero_name,
            defaults={
                "hero_type": HERO_TYPES.get(hero_name, "Universal"),
                "item1": specs[0]["cosmetic"] if len(specs) > 0 else None,
                "item2": specs[1]["cosmetic"] if len(specs) > 1 else None,
            },
        )
        if created:
            heroes_created += 1
        spec_ids = {spec["cosmetic"].id for spec in specs}
        stale_item = hero.item1 if hero.item1 and hero.item1.id not in spec_ids else None
        if stale_item:
            hero.item1 = specs[0]["cosmetic"]
            hero.save(update_fields=["item1"])

    return {
        "cosmetics_created": cosmetics_created,
        "heroes_created": heroes_created,
        "records_created": records_created,
        "hero_to_items": hero_to_items,
    }


def assign_holdings(trading, holdings, hero_to_items):
    assigned = []
    unresolved = []

    for slot_index in range(1, 11):
        setattr(trading, f"asset{slot_index}", None)

    for holding in holdings:
        options = hero_to_items.get(holding["hero_name"], [])
        selected = None

        if len(options) == 1:
            selected = options[0]["cosmetic"]
        elif len(options) > 1 and holding["current_price"] is not None:
            price_matches = [
                spec["cosmetic"]
                for spec in options
                if spec["latest_price"] == holding["current_price"]
            ]
            if len(price_matches) == 1:
                selected = price_matches[0]

        if selected is None and options:
            selected = options[0]["cosmetic"]

        if selected is None:
            unresolved.append(holding["hero_name"])
            continue

        assigned.append(selected)

    for slot_index, cosmetic in enumerate(assigned[:10], start=1):
        setattr(trading, f"asset{slot_index}", cosmetic)

    trading.save()
    return assigned, unresolved


class Command(BaseCommand):
    help = "Import DOTA2 workbook data into mocktrade models."

    def add_arguments(self, parser):
        parser.add_argument("--file", help="Path to the workbook. Defaults to dota2饰品.xlsx in project root.")
        parser.add_argument("--username", help="Bind holdings from the workbook to this Django auth user.")
        parser.add_argument("--password", help="Password used only if the auth user must be created.")
        parser.add_argument(
            "--replace",
            action="store_true",
            help="Clear Hero/Cosmetic/MarketRecord data before importing the workbook.",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        workbook_path = resolve_workbook_path(options["file"])
        workbook = load_workbook(workbook_path, data_only=True)
        item_specs = parse_database_sheet(workbook.worksheets[DATABASE_SHEET_INDEX])
        holdings = parse_holdings_sheet(workbook.worksheets[HOLDINGS_SHEET_INDEX])

        if not item_specs:
            raise CommandError("No market data rows found in the workbook.")

        if options["replace"]:
            Hero.objects.all().delete()
            Cosmetic.objects.all().delete()

        summary = import_market_data(item_specs)
        self.stdout.write(
            self.style.SUCCESS(
                f"Imported {len(item_specs)} cosmetics rows, "
                f"{sum(len(spec['market_points']) for spec in item_specs)} market points."
            )
        )
        self.stdout.write(
            f"Created cosmetics={summary['cosmetics_created']}, "
            f"heroes={summary['heroes_created']}, "
            f"market_records={summary['records_created']}."
        )

        username = options["username"]
        if not username:
            return

        _, auth_user_created = ensure_auth_user(username, options["password"])
        _, trading, user_account_created, trading_created = ensure_trading_account(username)
        assigned, unresolved = assign_holdings(trading, holdings, summary["hero_to_items"])

        self.stdout.write(
            self.style.SUCCESS(
                f"Linked workbook holdings to '{username}': "
                f"assigned={len(assigned)}, unresolved={len(unresolved)}."
            )
        )
        self.stdout.write(
            f"Auth user created={auth_user_created}, "
            f"mocktrade user created={user_account_created}, "
            f"trading account created={trading_created}."
        )

        if unresolved:
            self.stdout.write(self.style.WARNING(f"Unresolved holdings: {', '.join(unresolved)}"))
